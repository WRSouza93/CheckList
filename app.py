import io
import json
from datetime import datetime
from functools import wraps

from flask import (Flask, flash, jsonify, redirect, render_template,
                   request, send_file, session, url_for)

import database as db

app = Flask(__name__)
app.secret_key = "varpe-form25-secret-2025"

db.init_db()


# ═══════════════════════════════════════════════════════════════════════════
# Auth decorators
# ═══════════════════════════════════════════════════════════════════════════

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login", next=request.path))
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        if not session.get("is_admin"):
            flash("Acesso restrito a administradores.", "danger")
            return redirect(url_for("index"))
        return f(*args, **kwargs)
    return decorated


# ═══════════════════════════════════════════════════════════════════════════
# Auth routes
# ═══════════════════════════════════════════════════════════════════════════

@app.route("/login", methods=["GET", "POST"])
def login():
    if "user_id" in session:
        return redirect(url_for("index"))
    funcionarios = db.listar_funcionarios(apenas_ativos=True)
    if request.method == "POST":
        nome  = request.form.get("nome", "").strip()
        senha = request.form.get("senha", "")
        user  = db.verificar_login(nome, senha)
        if user:
            session["user_id"]   = user["id"]
            session["user_nome"] = user["nome"]
            session["is_admin"]  = bool(user["is_admin"])
            next_url = request.args.get("next") or url_for("index")
            return redirect(next_url)
        flash("Nome ou senha incorretos.", "danger")
    return render_template("login.html", funcionarios=funcionarios)


@app.route("/logout", methods=["POST"])
def logout():
    session.clear()
    return redirect(url_for("login"))


# ═══════════════════════════════════════════════════════════════════════════
# Helpers
# ═══════════════════════════════════════════════════════════════════════════

def _now():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _parse_respostas(form, existing_respostas, desc_map):
    respostas = dict(existing_respostas)
    item_codes = db.get_all_item_codes_db()
    user_id = session.get("user_id")
    now = _now()

    for codigo in item_codes:
        safe        = codigo.replace(".", "_")
        status      = form.get(f"status_{safe}", "").strip()
        resp_nome   = form.get(f"resp_{safe}", "").strip()
        if not session.get("is_admin") and status and not resp_nome:
            resp_nome = session.get("user_nome", "")
        incidencia  = form.get(f"inc_{safe}", "").strip()
        v_ideal     = form.get(f"vi_{safe}", "").strip()
        v_medido    = form.get(f"vm_{safe}", "").strip()
        equipamento = form.get(f"eq_{safe}", "").strip()

        existing = respostas.get(codigo, {})

        if not status and not resp_nome and not v_ideal and not v_medido and not equipamento:
            continue  # nada enviado — mantém

        if not status:
            if existing:
                existing.update({"responsavel": resp_nome, "valor_ideal": v_ideal,
                                  "valor_medido": v_medido, "equipamento": equipamento})
                respostas[codigo] = existing
            continue

        old_status     = existing.get("status", "")
        marcado_por_id = existing.get("marcado_por_id")
        historico      = list(existing.get("historico_nok", []))

        if old_status == "NOK" and status != "NOK" and old_status:
            historico.append({
                "status": "NOK", "responsavel": existing.get("responsavel", ""),
                "incidencia": existing.get("incidencia", ""),
                "data": existing.get("marcado_em", now),
                "acao": "resolved", "resolvido_por_id": user_id,
            })
        elif old_status == "NOK" and status == "NOK":
            historico.append({
                "status": "NOK", "responsavel": existing.get("responsavel", ""),
                "incidencia": existing.get("incidencia", ""),
                "data": existing.get("marcado_em", now),
                "acao": "reeval_nok", "avaliado_por_id": user_id,
            })

        if not marcado_por_id:
            marcado_por_id = user_id

        respostas[codigo] = {
            "status":       status,
            "responsavel":  resp_nome,
            "incidencia":   incidencia,
            "valor_ideal":  v_ideal,
            "valor_medido": v_medido,
            "equipamento":  equipamento,
            "marcado_por_id": marcado_por_id,
            "marcado_em":   existing.get("marcado_em", now) if old_status else now,
            "desc_snapshot": existing.get("desc_snapshot") or desc_map.get(codigo, ""),
            "historico_nok": historico,
        }
    return respostas


def _contagem(respostas):
    ok  = sum(1 for v in respostas.values() if v.get("status") == "OK")
    nok = sum(1 for v in respostas.values() if v.get("status") == "NOK")
    na  = sum(1 for v in respostas.values() if v.get("status") == "NA")
    return ok, nok, na


def _todos_preenchidos(respostas, total_items):
    ok, nok, na = _contagem(respostas)
    return (ok + nok + na) >= total_items


# ═══════════════════════════════════════════════════════════════════════════
# Registros — CRUD
# ═══════════════════════════════════════════════════════════════════════════

@app.route("/")
@login_required
def index():
    busca   = request.args.get("q", "").strip()
    filtro  = request.args.get("f", "todos")
    registros = db.listar_registros(busca or None)
    if filtro == "abertos":
        registros = [r for r in registros if not r.get("encerrado")]
    elif filtro == "encerrados":
        registros = [r for r in registros if r.get("encerrado")]
    return render_template("index.html", registros=registros, busca=busca, filtro=filtro)


@app.route("/novo", methods=["GET", "POST"])
@admin_required
def novo():
    checklist    = db.listar_checklist_db()
    funcionarios = db.listar_funcionarios(apenas_ativos=True)
    if request.method == "POST":
        dados = {k: request.form.get(k, "") for k in
                 ("cliente", "op", "ns", "modelo", "data", "producao", "responsavel")}
        desc_map  = db.get_item_desc_map()
        respostas = _parse_respostas(request.form, {}, desc_map)
        nok_sem_inc = [c for c, v in respostas.items()
                       if v.get("status") == "NOK" and not v.get("incidencia", "").strip()]
        if nok_sem_inc:
            flash(f"Preencha a incidência dos itens NOK: {', '.join(nok_sem_inc)}", "danger")
            return render_template("form.html", registro=None, checklist=checklist,
                                   funcionarios=funcionarios, respostas=respostas,
                                   nok_sem_inc=nok_sem_inc)
        novo_id = db.criar_registro(dados, respostas, criado_por_id=session["user_id"])
        flash("Registro criado com sucesso!", "success")
        return redirect(url_for("ver", registro_id=novo_id))
    return render_template("form.html", registro=None, checklist=checklist,
                           funcionarios=funcionarios, respostas={}, nok_sem_inc=[])


@app.route("/ver/<int:registro_id>")
@login_required
def ver(registro_id):
    registro = db.obter_registro(registro_id)
    if not registro:
        flash("Registro não encontrado.", "danger")
        return redirect(url_for("index"))
    checklist = db.listar_checklist_db()
    total     = len(db.get_all_item_codes_db())
    ok, nok, na = _contagem(registro["respostas"])
    funcs = {f["id"]: f["nome"] for f in db.listar_funcionarios()}
    return render_template("view.html", registro=registro, checklist=checklist,
                           ok=ok, nok=nok, na=na, total=total,
                           todos_preenchidos=_todos_preenchidos(registro["respostas"], total),
                           funcs=funcs)


@app.route("/editar/<int:registro_id>", methods=["GET", "POST"])
@login_required
def editar(registro_id):
    registro = db.obter_registro(registro_id)
    if not registro:
        flash("Registro não encontrado.", "danger")
        return redirect(url_for("index"))
    if registro.get("encerrado") and not session.get("is_admin"):
        flash("Registro encerrado. Apenas administradores podem editar.", "warning")
        return redirect(url_for("ver", registro_id=registro_id))

    checklist    = db.listar_checklist_db()
    funcionarios = db.listar_funcionarios(apenas_ativos=True)

    if request.method == "POST":
        if session.get("is_admin"):
            dados = {k: request.form.get(k, "") for k in
                     ("cliente", "op", "ns", "modelo", "data", "producao", "responsavel")}
        else:
            dados = {k: registro.get(k, "") for k in
                     ("cliente", "op", "ns", "modelo", "data", "producao", "responsavel")}

        desc_map  = db.get_item_desc_map()
        respostas = _parse_respostas(request.form, registro["respostas"], desc_map)

        nok_sem_inc = [c for c, v in respostas.items()
                       if v.get("status") == "NOK" and not v.get("incidencia", "").strip()]
        if nok_sem_inc:
            flash(f"Preencha a incidência dos itens NOK: {', '.join(nok_sem_inc)}", "danger")
            return render_template("form.html", registro=registro, checklist=checklist,
                                   funcionarios=funcionarios, respostas=respostas,
                                   nok_sem_inc=nok_sem_inc)

        db.atualizar_registro(registro_id, dados, respostas)
        flash("Registro atualizado com sucesso!", "success")
        return redirect(url_for("ver", registro_id=registro_id))

    return render_template("form.html", registro=registro, checklist=checklist,
                           funcionarios=funcionarios, respostas=registro["respostas"],
                           nok_sem_inc=[])


@app.route("/excluir/<int:registro_id>", methods=["POST"])
@admin_required
def excluir(registro_id):
    db.excluir_registro(registro_id)
    flash("Registro excluído.", "info")
    return redirect(url_for("index"))


# ── Limpar item (AJAX) ────────────────────────────────────────────────────

@app.route("/registro/<int:registro_id>/limpar_item", methods=["POST"])
@login_required
def limpar_item(registro_id):
    codigo   = (request.json or {}).get("codigo", "")
    registro = db.obter_registro(registro_id)
    if not registro:
        return jsonify({"ok": False, "msg": "Não encontrado"}), 404
    resp         = registro["respostas"].get(codigo, {})
    marcado_por  = resp.get("marcado_por_id")
    if marcado_por and not session.get("is_admin") and marcado_por != session.get("user_id"):
        return jsonify({"ok": False, "msg": "Sem permissão para desmarcar este item."}), 403
    db.limpar_item_resposta(registro_id, codigo)
    return jsonify({"ok": True})


# ── Encerrar / Reabrir ────────────────────────────────────────────────────

@app.route("/registro/<int:registro_id>/encerrar", methods=["POST"])
@admin_required
def encerrar(registro_id):
    registro = db.obter_registro(registro_id)
    if not registro:
        flash("Registro não encontrado.", "danger")
        return redirect(url_for("index"))
    total = len(db.get_all_item_codes_db())
    if not _todos_preenchidos(registro["respostas"], total):
        flash("Todos os itens devem estar preenchidos antes de encerrar.", "warning")
        return redirect(url_for("ver", registro_id=registro_id))
    assin_q = request.form.get("assinatura_qualidade", "")
    assin_p = request.form.get("assinatura_producao", "")
    if not assin_q or not assin_p:
        flash("Ambas as assinaturas são obrigatórias.", "warning")
        return redirect(url_for("ver", registro_id=registro_id))
    db.encerrar_registro(registro_id, assin_q, assin_p)
    flash("Checklist encerrada com sucesso!", "success")
    return redirect(url_for("ver", registro_id=registro_id))


@app.route("/registro/<int:registro_id>/reabrir", methods=["POST"])
@admin_required
def reabrir(registro_id):
    db.reabrir_registro(registro_id)
    flash("Checklist reaberta.", "info")
    return redirect(url_for("ver", registro_id=registro_id))


# ═══════════════════════════════════════════════════════════════════════════
# Funcionários / Usuários
# ═══════════════════════════════════════════════════════════════════════════

@app.route("/funcionarios")
@login_required
def funcionarios():
    lista = db.listar_funcionarios()
    return render_template("funcionarios.html", funcionarios=lista)


@app.route("/funcionarios/novo", methods=["GET", "POST"])
@admin_required
def funcionario_novo():
    if request.method == "POST":
        nome     = request.form.get("nome", "").strip()
        senha    = request.form.get("senha", "").strip() or None
        is_admin = request.form.get("is_admin") == "1"
        if not nome:
            flash("Nome é obrigatório.", "danger")
        else:
            ok, erro = db.criar_funcionario(nome, senha, is_admin)
            if ok:
                flash(f"Funcionário '{nome}' cadastrado!", "success")
                return redirect(url_for("funcionarios"))
            flash(erro, "danger")
    return render_template("funcionarios.html", funcionarios=db.listar_funcionarios(), modo="novo")


@app.route("/funcionarios/editar/<int:func_id>", methods=["GET", "POST"])
@login_required
def funcionario_editar(func_id):
    func = db.obter_funcionario(func_id)
    if not func:
        flash("Funcionário não encontrado.", "danger")
        return redirect(url_for("funcionarios"))
    if not session.get("is_admin") and session["user_id"] != func_id:
        flash("Sem permissão.", "danger")
        return redirect(url_for("funcionarios"))

    if request.method == "POST":
        nome     = request.form.get("nome", "").strip()
        ativo    = request.form.get("ativo", "1") == "1"
        senha    = request.form.get("senha", "").strip() or None
        is_admin = (request.form.get("is_admin") == "1") if session.get("is_admin") else bool(func["is_admin"])
        if not nome:
            flash("Nome é obrigatório.", "danger")
        else:
            ok, erro = db.atualizar_funcionario(func_id, nome, ativo, is_admin=is_admin, senha_plain=senha)
            if ok:
                if session["user_id"] == func_id:
                    session["user_nome"] = nome
                flash("Funcionário atualizado!", "success")
                return redirect(url_for("funcionarios"))
            flash(erro, "danger")

    return render_template("funcionarios.html", funcionarios=db.listar_funcionarios(),
                           modo="editar", editando=func)


@app.route("/funcionarios/excluir/<int:func_id>", methods=["POST"])
@admin_required
def funcionario_excluir(func_id):
    if func_id == session.get("user_id"):
        flash("Você não pode excluir sua própria conta.", "danger")
        return redirect(url_for("funcionarios"))
    if db.funcionario_em_uso(func_id):
        db.inativar_funcionario(func_id)
        flash("Funcionário inativado (possui registros associados).", "warning")
    else:
        db.excluir_funcionario(func_id)
        flash("Funcionário excluído.", "info")
    return redirect(url_for("funcionarios"))


# ═══════════════════════════════════════════════════════════════════════════
# Checklist — Administração
# ═══════════════════════════════════════════════════════════════════════════

@app.route("/checklist-admin")
@admin_required
def checklist_admin():
    secoes = db.listar_secoes()
    secoes_com_dados = []
    for s in secoes:
        subs        = db.listar_subsecoes(secao_id=s["id"])
        itens_dir   = db.listar_itens_admin(secao_id=s["id"]) if not subs else []
        for sub in subs:
            sub["itens"] = db.listar_itens_admin(subsecao_id=sub["id"])
        secoes_com_dados.append({**s, "subsecoes": subs, "itens_diretos": itens_dir})
    todas_secoes = db.listar_secoes(apenas_ativas=True)
    return render_template("checklist_admin.html",
                           secoes=secoes_com_dados, todas_secoes=todas_secoes)


@app.route("/checklist-admin/secao/nova", methods=["POST"])
@admin_required
def checklist_secao_nova():
    codigo = request.form.get("codigo", "").strip()
    nome   = request.form.get("nome", "").strip()
    ordem  = int(request.form.get("ordem", 99))
    if not codigo or not nome:
        flash("Código e nome são obrigatórios.", "danger")
    else:
        ok, erro = db.criar_secao(codigo, nome, ordem)
        flash("Seção criada!" if ok else erro, "success" if ok else "danger")
    return redirect(url_for("checklist_admin"))


@app.route("/checklist-admin/secao/<int:secao_id>/editar", methods=["POST"])
@admin_required
def checklist_secao_editar(secao_id):
    db.atualizar_secao(secao_id,
                       request.form.get("nome", "").strip(),
                       int(request.form.get("ordem", 0)),
                       request.form.get("ativo", "1") == "1")
    flash("Seção atualizada!", "success")
    return redirect(url_for("checklist_admin"))


@app.route("/checklist-admin/subsecao/nova", methods=["POST"])
@admin_required
def checklist_subsecao_nova():
    secao_id = int(request.form.get("secao_id", 0))
    codigo   = request.form.get("codigo", "").strip()
    nome     = request.form.get("nome", "").strip()
    ordem    = int(request.form.get("ordem", 99))
    if not codigo or not nome:
        flash("Código e nome são obrigatórios.", "danger")
    else:
        ok, erro = db.criar_subsecao(secao_id, codigo, nome, ordem)
        flash("Subseção criada!" if ok else erro, "success" if ok else "danger")
    return redirect(url_for("checklist_admin"))


@app.route("/checklist-admin/subsecao/<int:sub_id>/editar", methods=["POST"])
@admin_required
def checklist_subsecao_editar(sub_id):
    db.atualizar_subsecao(sub_id,
                          request.form.get("nome", "").strip(),
                          int(request.form.get("ordem", 0)),
                          request.form.get("ativo", "1") == "1")
    flash("Subseção atualizada!", "success")
    return redirect(url_for("checklist_admin"))


@app.route("/checklist-admin/item/novo", methods=["POST"])
@admin_required
def checklist_item_novo():
    secao_id    = request.form.get("secao_id") or None
    subsecao_id = request.form.get("subsecao_id") or None
    codigo      = request.form.get("codigo", "").strip()
    descricao   = request.form.get("descricao", "").strip()
    tem_medicao = request.form.get("tem_medicao") == "1"
    ordem       = int(request.form.get("ordem", 99))
    if not codigo or not descricao:
        flash("Código e descrição são obrigatórios.", "danger")
    else:
        ok, erro = db.criar_item_checklist(secao_id, subsecao_id, codigo, descricao, tem_medicao, ordem)
        flash("Item criado!" if ok else erro, "success" if ok else "danger")
    return redirect(url_for("checklist_admin"))


@app.route("/checklist-admin/item/<int:item_id>/editar", methods=["POST"])
@admin_required
def checklist_item_editar(item_id):
    item = db.obter_item_checklist(item_id)
    if not item:
        flash("Item não encontrado.", "danger")
        return redirect(url_for("checklist_admin"))
    ativo_val = request.form.get("ativo", "1") == "1"
    if db.item_checklist_em_uso(item["codigo"]) and not ativo_val:
        flash("Item em uso em registros existentes — foi inativado.", "warning")
    db.atualizar_item_checklist(item_id,
                                request.form.get("descricao", "").strip(),
                                request.form.get("tem_medicao") == "1",
                                int(request.form.get("ordem", item["ordem"])),
                                ativo_val)
    flash("Item atualizado!", "success")
    return redirect(url_for("checklist_admin"))


# ═══════════════════════════════════════════════════════════════════════════
# Export — Excel
# ═══════════════════════════════════════════════════════════════════════════

@app.route("/exportar/excel/<int:registro_id>")
@login_required
def exportar_excel(registro_id):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    registro = db.obter_registro(registro_id)
    if not registro:
        flash("Registro não encontrado.", "danger")
        return redirect(url_for("index"))

    checklist = db.listar_checklist_db()
    wb = Workbook()
    ws = wb.active
    ws.title = "Checklist"
    AZUL, CINZA, VERDE, VERMELHO, AMARELO = "082D45","F2F4F6","D4EDDA","F8D7DA","FFF3CD"

    def cs(cell, bold=False, bg=None, align="left", fc="000000", size=10):
        cell.font = Font(bold=bold, color=fc, size=size)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        if bg: cell.fill = PatternFill("solid", fgColor=bg)
        thin = Side(style="thin", color="CCCCCC")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:H1")
    ws["A1"] = "FORM 25 - INSPEÇÃO FINAL - VARPE"
    cs(ws["A1"], bold=True, bg=AZUL, align="center", fc="FFFFFF", size=14)
    ws.row_dimensions[1].height = 28

    for i, (lbl, val) in enumerate([("Cliente",registro["cliente"]),("OP",registro["op"]),
            ("N/S",registro["ns"]),("Modelo",registro["modelo"]),("Data",registro["data"]),
            ("Produção",registro["producao"]),("Responsável",registro["responsavel"])], 2):
        ws.cell(row=i,column=1,value=lbl); cs(ws.cell(row=i,column=1),bold=True,bg=CINZA)
        ws.merge_cells(start_row=i,start_column=2,end_row=i,end_column=8)
        ws.cell(row=i,column=2,value=val); cs(ws.cell(row=i,column=2))

    row = 10
    for col, h in enumerate(["#","Item","Status","Responsável","Incidência","V. Ideal","V. Medido","Equipamento"],1):
        ws.cell(row=row,column=col,value=h); cs(ws.cell(row=row,column=col),bold=True,bg=AZUL,align="center",fc="FFFFFF")
    ws.row_dimensions[row].height = 18; row += 1

    respostas = registro["respostas"]
    for secao in checklist:
        ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=8)
        ws.cell(row=row,column=1,value=f"{secao['id']}. {secao['nome']}")
        cs(ws.cell(row=row,column=1),bold=True,bg="1A4A6E",fc="FFFFFF",size=11)
        ws.row_dimensions[row].height=20; row+=1

        items_flat = []
        if "subsecoes" in secao:
            for sub in secao["subsecoes"]:
                items_flat.append(("sub",sub["id"],sub["nome"],None))
                for it in sub["itens"]: items_flat.append(("item",it["id"],it["desc"],it))
        else:
            for it in secao.get("itens",[]): items_flat.append(("item",it["id"],it["desc"],it))

        for tipo, iid, desc, it_data in items_flat:
            if tipo=="sub":
                ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=8)
                ws.cell(row=row,column=1,value=f"  {iid} - {desc}")
                cs(ws.cell(row=row,column=1),bold=True,bg="D0E8F5"); ws.row_dimensions[row].height=16; row+=1; continue
            resp=respostas.get(iid,{}); s=resp.get("status","")
            bg=VERDE if s=="OK" else VERMELHO if s=="NOK" else AMARELO if s=="NA" else None
            desc_show = resp.get("desc_snapshot") or desc
            for col,val in enumerate([iid,desc_show,s,resp.get("responsavel",""),
                resp.get("incidencia",""),resp.get("valor_ideal",""),resp.get("valor_medido",""),resp.get("equipamento","")],1):
                ws.cell(row=row,column=col,value=val)
                cs(ws.cell(row=row,column=col),bg=bg,align="center" if col in(1,3,4,6,7) else "left",bold=(col==3))
            ws.row_dimensions[row].height=15; row+=1

    for col,w in zip(["A","B","C","D","E","F","G","H"],[10,55,10,14,30,14,14,22]):
        ws.column_dimensions[col].width=w

    out=io.BytesIO(); wb.save(out); out.seek(0)
    return send_file(out,mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True,download_name=f"form25_{registro['op'] or registro_id}.xlsx")


@app.route("/exportar/excel")
@login_required
def exportar_excel_lista():
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    registros=db.listar_registros(); wb=Workbook(); ws=wb.active; ws.title="Registros"
    for col,h in enumerate(["ID","Cliente","OP","N/S","Modelo","Data","Produção","Responsável","Status","Criado em"],1):
        c=ws.cell(row=1,column=col,value=h); c.font=Font(bold=True,color="FFFFFF")
        c.fill=PatternFill("solid",fgColor="082D45"); c.alignment=Alignment(horizontal="center",vertical="center")
    for i,reg in enumerate(registros,2):
        for col,key in enumerate(["id","cliente","op","ns","modelo","data","producao","responsavel"],1):
            ws.cell(row=i,column=col,value=reg.get(key,""))
        ws.cell(row=i,column=9,value="Encerrado" if reg.get("encerrado") else "Em andamento")
        ws.cell(row=i,column=10,value=reg.get("created_at",""))
    for col in ["A","B","C","D","E","F","G","H","I","J"]: ws.column_dimensions[col].width=18
    out=io.BytesIO(); wb.save(out); out.seek(0)
    return send_file(out,mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True,download_name="form25_lista.xlsx")


# ═══════════════════════════════════════════════════════════════════════════
# Export — PDF
# ═══════════════════════════════════════════════════════════════════════════

@app.route("/exportar/pdf/<int:registro_id>")
@login_required
def exportar_pdf(registro_id):
    import base64
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import Image, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    registro=db.obter_registro(registro_id)
    if not registro: flash("Não encontrado.","danger"); return redirect(url_for("index"))
    checklist=db.listar_checklist_db()
    out=io.BytesIO()
    doc=SimpleDocTemplate(out,pagesize=A4,leftMargin=1.5*cm,rightMargin=1.5*cm,topMargin=2*cm,bottomMargin=2*cm)
    cp=colors.HexColor("#082D45"); cs=colors.HexColor("#1A4A6E")
    csub=colors.HexColor("#D0E8F5"); cok=colors.HexColor("#D4EDDA"); cnok=colors.HexColor("#F8D7DA"); cna=colors.HexColor("#FFF3CD")
    st=lambda n,**k: ParagraphStyle(n,**k)
    st_t=st("t",fontSize=14,textColor=colors.white,fontName="Helvetica-Bold",alignment=1)
    st_iL=st("iL",fontSize=9,fontName="Helvetica-Bold"); st_iV=st("iV",fontSize=9,fontName="Helvetica")
    st_sec=st("sec",fontSize=10,textColor=colors.white,fontName="Helvetica-Bold")
    st_sub=st("sub",fontSize=9,textColor=cs,fontName="Helvetica-Bold")
    st_item=st("item",fontSize=8,fontName="Helvetica",leading=10)
    st_ok=st("ok",fontSize=8,fontName="Helvetica-Bold",textColor=colors.HexColor("#155724"),alignment=1)
    st_nok=st("nok",fontSize=8,fontName="Helvetica-Bold",textColor=colors.HexColor("#721C24"),alignment=1)
    st_na=st("na",fontSize=8,fontName="Helvetica-Bold",textColor=colors.HexColor("#856404"),alignment=1)
    st_vz=st("vz",fontSize=8,fontName="Helvetica",textColor=colors.grey,alignment=1)
    story=[]
    tt=Table([[Paragraph("FORM 25 — INSPEÇÃO FINAL — VARPE",st_t)]],colWidths=[18*cm])
    tt.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),cp),("TOPPADDING",(0,0),(-1,-1),8),("BOTTOMPADDING",(0,0),(-1,-1),8)]))
    story.append(tt); story.append(Spacer(1,0.4*cm))
    info_rows=[
        [Paragraph("Cliente",st_iL),Paragraph(registro["cliente"] or "-",st_iV),Paragraph("OP",st_iL),Paragraph(registro["op"] or "-",st_iV)],
        [Paragraph("N/S",st_iL),Paragraph(registro["ns"] or "-",st_iV),Paragraph("Modelo",st_iL),Paragraph(registro["modelo"] or "-",st_iV)],
        [Paragraph("Data",st_iL),Paragraph(registro["data"] or "-",st_iV),Paragraph("Produção",st_iL),Paragraph(registro["producao"] or "-",st_iV)],
        [Paragraph("Responsável",st_iL),Paragraph(registro["responsavel"] or "-",st_iV),"",""],
    ]
    info_t=Table(info_rows,colWidths=[3*cm,5.5*cm,3*cm,6.5*cm])
    info_t.setStyle(TableStyle([("GRID",(0,0),(-1,-1),0.5,colors.HexColor("#CCCCCC")),
        ("BACKGROUND",(0,0),(0,-1),colors.HexColor("#F2F4F6")),("BACKGROUND",(2,0),(2,-1),colors.HexColor("#F2F4F6")),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),("LEFTPADDING",(0,0),(-1,-1),5)]))
    story.append(info_t); story.append(Spacer(1,0.5*cm))
    respostas=registro["respostas"]; ok,nok,na=_contagem(respostas); total=len(db.get_all_item_codes_db())
    resumo=Table([[
        Paragraph(f"✓ OK: {ok}",st("rOK",fontSize=9,fontName="Helvetica-Bold",textColor=colors.HexColor("#155724"),alignment=1)),
        Paragraph(f"✗ NOK: {nok}",st("rNOK",fontSize=9,fontName="Helvetica-Bold",textColor=colors.HexColor("#721C24"),alignment=1)),
        Paragraph(f"— NA: {na}",st("rNA",fontSize=9,fontName="Helvetica-Bold",textColor=colors.HexColor("#856404"),alignment=1)),
        Paragraph(f"Total: {total}",st("rT",fontSize=9,fontName="Helvetica-Bold",alignment=1)),
    ]],colWidths=[4.5*cm]*4)
    resumo.setStyle(TableStyle([("BACKGROUND",(0,0),(0,-1),cok),("BACKGROUND",(1,0),(1,-1),cnok),
        ("BACKGROUND",(2,0),(2,-1),cna),("BACKGROUND",(3,0),(3,-1),colors.HexColor("#E9ECEF")),
        ("GRID",(0,0),(-1,-1),0.5,colors.HexColor("#CCCCCC")),("TOPPADDING",(0,0),(-1,-1),6),("BOTTOMPADDING",(0,0),(-1,-1),6)]))
    story.append(resumo); story.append(Spacer(1,0.5*cm))
    for secao in checklist:
        sec_t=Table([[Paragraph(f"{secao['id']}. {secao['nome']}",st_sec)]],colWidths=[18*cm])
        sec_t.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),cs),("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5),("LEFTPADDING",(0,0),(-1,-1),6)]))
        story.append(sec_t)
        def make_rows(items_list):
            rows=[]
            for it in items_list:
                resp=respostas.get(it["id"],{}); s=resp.get("status","")
                desc_show=resp.get("desc_snapshot") or it["desc"]
                st_s=st_ok if s=="OK" else st_nok if s=="NOK" else st_na if s=="NA" else st_vz
                rows.append([Paragraph(it["id"],st("iid",fontSize=7,fontName="Helvetica",alignment=1)),
                    Paragraph(desc_show,st_item),Paragraph(s or "—",st_s),
                    Paragraph(resp.get("responsavel",""),st("rp",fontSize=8,fontName="Helvetica",alignment=1)),
                    Paragraph(resp.get("incidencia",""),st_item)])
            return rows
        def apply_colors(t,items_list):
            for i,it in enumerate(items_list):
                s=respostas.get(it["id"],{}).get("status","")
                bg=cok if s=="OK" else cnok if s=="NOK" else cna if s=="NA" else None
                if bg: t.setStyle(TableStyle([("BACKGROUND",(2,i),(2,i),bg)]))
        tstyle=TableStyle([("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#DDDDDD")),
            ("VALIGN",(0,0),(-1,-1),"MIDDLE"),("TOPPADDING",(0,0),(-1,-1),2),
            ("BOTTOMPADDING",(0,0),(-1,-1),2),("LEFTPADDING",(0,0),(-1,-1),4)])
        if "subsecoes" in secao:
            for sub in secao["subsecoes"]:
                sub_t=Table([[Paragraph(f"  {sub['id']} — {sub['nome']}",st_sub)]],colWidths=[18*cm])
                sub_t.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),csub),("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),("LEFTPADDING",(0,0),(-1,-1),6)]))
                story.append(sub_t)
                rows=make_rows(sub["itens"])
                if rows:
                    t=Table(rows,colWidths=[1.6*cm,7.5*cm,1.4*cm,2.2*cm,5.3*cm]); t.setStyle(tstyle); apply_colors(t,sub["itens"]); story.append(t)
        else:
            rows=make_rows(secao.get("itens",[]))
            if rows:
                t=Table(rows,colWidths=[1.6*cm,7.5*cm,1.4*cm,2.2*cm,5.3*cm]); t.setStyle(tstyle); apply_colors(t,secao.get("itens",[])); story.append(t)
        story.append(Spacer(1,0.3*cm))

    # ── Página de assinaturas ────────────────────────────────────────────────
    if registro.get("encerrado") and (registro.get("assinatura_qualidade") or registro.get("assinatura_producao")):
        story.append(PageBreak())
        # Título da seção
        assin_header = Table([[Paragraph("ASSINATURAS", st("ah", fontSize=12, textColor=colors.white, fontName="Helvetica-Bold", alignment=1))]], colWidths=[18*cm])
        assin_header.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),cp),("TOPPADDING",(0,0),(-1,-1),8),("BOTTOMPADDING",(0,0),(-1,-1),8)]))
        story.append(assin_header)
        story.append(Spacer(1, 0.6*cm))

        encerrado_em = registro.get("encerrado_em") or registro.get("updated_at", "")
        # Formata data legível
        try:
            dt = datetime.strptime(encerrado_em[:19], "%Y-%m-%d %H:%M:%S")
            data_fmt = dt.strftime("%d/%m/%Y %H:%M")
        except Exception:
            data_fmt = encerrado_em or "—"

        st_label = st("sl", fontSize=9, fontName="Helvetica-Bold", textColor=cs, alignment=1)
        st_date  = st("sd", fontSize=8, fontName="Helvetica", textColor=colors.grey, alignment=1)

        def _sig_img(b64_data, w=7.5*cm, h=3.5*cm):
            if not b64_data:
                return Paragraph("(sem assinatura)", st("ns", fontSize=8, fontName="Helvetica", textColor=colors.grey, alignment=1))
            try:
                raw = b64_data.split(",", 1)[-1]
                img_bytes = base64.b64decode(raw)
                return Image(io.BytesIO(img_bytes), width=w, height=h)
            except Exception:
                return Paragraph("(erro ao carregar assinatura)", st("er", fontSize=8, fontName="Helvetica", textColor=colors.red, alignment=1))

        sig_q = _sig_img(registro.get("assinatura_qualidade"))
        sig_p = _sig_img(registro.get("assinatura_producao"))

        assin_table = Table([
            [Paragraph("Qualidade", st_label), Paragraph("Produção", st_label)],
            [sig_q, sig_p],
            [Paragraph(f"Data: {data_fmt}", st_date), Paragraph(f"Data: {data_fmt}", st_date)],
        ], colWidths=[9*cm, 9*cm])
        assin_table.setStyle(TableStyle([
            ("GRID",     (0,0), (-1,-1), 0.5, colors.HexColor("#CCCCCC")),
            ("VALIGN",   (0,0), (-1,-1), "MIDDLE"),
            ("ALIGN",    (0,0), (-1,-1), "CENTER"),
            ("TOPPADDING",    (0,0), (-1,-1), 8),
            ("BOTTOMPADDING", (0,0), (-1,-1), 8),
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#F2F4F6")),
        ]))
        story.append(assin_table)

    doc.build(story); out.seek(0)
    return send_file(out,mimetype="application/pdf",as_attachment=True,download_name=f"form25_{registro['op'] or registro_id}.pdf")


@app.route("/exportar/pdf")
@login_required
def exportar_pdf_lista():
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    registros=db.listar_registros(); out=io.BytesIO()
    doc=SimpleDocTemplate(out,pagesize=A4,leftMargin=1.5*cm,rightMargin=1.5*cm,topMargin=2*cm,bottomMargin=2*cm)
    cp=colors.HexColor("#082D45"); st_t=ParagraphStyle("t",fontSize=14,textColor=colors.white,fontName="Helvetica-Bold",alignment=1)
    st_h=ParagraphStyle("h",fontSize=9,fontName="Helvetica-Bold",textColor=colors.white,alignment=1)
    st_c=ParagraphStyle("c",fontSize=8,fontName="Helvetica"); story=[]
    tt=Table([[Paragraph("FORM 25 — LISTA DE REGISTROS — VARPE",st_t)]],colWidths=[18*cm])
    tt.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),cp),("TOPPADDING",(0,0),(-1,-1),8),("BOTTOMPADDING",(0,0),(-1,-1),8)]))
    story.append(tt); story.append(Spacer(1,0.5*cm))
    rows=[[Paragraph(h,st_h) for h in ["ID","Cliente","OP","N/S","Modelo","Data","Responsável","Status"]]]
    for reg in registros:
        rows.append([Paragraph(str(reg["id"]),st_c),Paragraph(reg["cliente"] or "",st_c),
            Paragraph(reg["op"] or "",st_c),Paragraph(reg["ns"] or "",st_c),Paragraph(reg["modelo"] or "",st_c),
            Paragraph(reg["data"] or "",st_c),Paragraph(reg["responsavel"] or "",st_c),
            Paragraph("Encerrado" if reg.get("encerrado") else "Em andamento",st_c)])
    t=Table(rows,colWidths=[1*cm,3*cm,2.5*cm,2.5*cm,3*cm,2.5*cm,2.5*cm,2*cm])
    t.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),cp),("GRID",(0,0),(-1,-1),0.5,colors.HexColor("#CCCCCC")),
        ("ROWBACKGROUND",(0,1),(-1,-1),[colors.white,colors.HexColor("#F8F9FA")]),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),("LEFTPADDING",(0,0),(-1,-1),5)]))
    story.append(t); doc.build(story); out.seek(0)
    return send_file(out,mimetype="application/pdf",as_attachment=True,download_name="form25_lista.pdf")


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
